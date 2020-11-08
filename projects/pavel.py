from openpyxl import load_workbook as lw
from datetime import datetime

lesson_rows = (2, 5, 7, 9, 11, 13, 15)
file = lw('расписание.xlsx')
page = file.active


def main():
    print('Начало работы')

    while True:
        answer = listen()
        if answer == 'exit':
            break
        elif answer == 'Расписание':
            schedule = get_schedule()


def listen():
    return input('Введите команду. Варианты:\nРасписание\nДЗ\nСоздать ДЗ\n')


def get_schedule():
    """
    Функция Расписание.
    Варианты вывода: на сегодня, на завтра, на неделю
    :return:
    """
    weekdays = {
        0: 2,    # Понедельник
        1: 5,    # Вторник
        2: 8,    # Среда
        3: 11,   # Четверг
        4: 14    # Пятница
    }
    today_weekday = datetime.weekday(datetime.today())
    today_weekday = weekdays[today_weekday]

    answer = input('______________\nВыберите один из вариантов:\nНа сегодня,\nНа завтра,\nНа неделю: ')
    if answer == 'На сегодня':
        return get_schedule_today(today_weekday)
    elif answer == 'На завтра':
        return get_schedule_tomorrow()
    elif answer == 'На неделю':
        return get_schedule_week()
    # elif answer == 'Назад':
    else:
        print('Ошибка ввода')


def get_schedule_today(weekday_colummn):
    """
    Выводит из файла расписание.xlsx расписание на текущий день
    :return:
    """
    print('__________________')
    for row in lesson_rows:
        print(page.cell(row=row, column=weekday_colummn).value)
        for i in range(1, 3):
            if page.cell(row=row, column=weekday_colummn + i).value is not None:
                print(page.cell(row=row, column=weekday_colummn + i).value)


def get_schedule_tomorrow():
    return 'Завтра будет МЯСО'


def get_schedule_week():
    return 'На неделе будет МЯСО'


if __name__ == '__main__':
    main()
