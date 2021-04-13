
from openpyxl import load_workbook
# from openpyxl import load_workbook as lw TODO заменить
# TODO продумать как будут передаваться определенные заявки опредленным "исполнителям"
# TODO сделать команду для замены номера телефона испонителя
# TODO пересмотреть статус заявки отправлено/просмотрено/выполненено...
# TODO подобаблять команды новые Юmy_status_id


def main():
    print('Начало работы')

# def find_empty_line

    while True:
        answer = listen()

        if answer == 'help':
            print(show_help())
        elif answer == 'reg_task':
            print('Регистрация заявки началась')
            reg_task()
        elif answer == 'my_status_id':
            id = input('Введите id: ')
            my_status_id(id)
        elif answer == 'status':
            status()
            # print(wb.sheetnames)
            # for i in range(1, 11):
            #     for j in range(1, 5):
            #         print(sheet.cell(row=i, column=j).value, end='\t')
            #     print()
        else:
            print('Ошибка, проверьте список команд введя "help"')


def listen():
    print('\n----------------------------------------------'
          ' \nВведите команду(help, reg_task, status):'
          '  \n---------------------------------------------- ')

    return input()


def show_help():
    return'''help - команда, которая выводит описание всех остальных команд
    reg_task - запускает процесс заполнения заявки
    status - показывает все имеющиеся заявки
    my_status_id - показывает статус заявки'''


def reg_task():
    global line
    role = input('Вы ученик или учитель? ')
    task = input('Опишите вашу проблему. Что нужно сделать? ')
    user_id = input('id ')   # тоже через бота
    name = input('Введите ваши ФИО ')    # потом через бота
    number = input('Номер телефона по которому с вами можно связаться: ')
    # сделать проверку ввода пользователя
    # Не нужно делать так вся информация нуждающаяся в проверке будет браться автоматически из телеги
    # number = int(input('Номер телефона по которому с вами можно связаться: ')) # это проверка собственно
    sheet.cell(row=line, column=1).value = role
    sheet.cell(row=line, column=2).value = task
    sheet.cell(row=line, column=3).value = user_id
    sheet.cell(row=line, column=4).value = name
    sheet.cell(row=line, column=5).value = number
    sheet.cell(row=line, column=6).value = 'Не выполнено'
    #   TODO продумать как будет меняться на 'Выполнено', получается нужно написать команду для "исполнителей"
    #   сделать к фестивалю проектов: прененосить выполненные заявки на второй лист, а ячейку на первом листе очищать
    wb.save('proverochka.xlsx')
    line += 1
    print('Ваша завка была принята.')


def print_task(row):
    for j in range(1, 6):
        print(sheet.cell(row=row, column=j).value, end='\t')


def status():
    print(wb.sheetnames)
    print(f'{line} записей:')
    for i in range(1, line):
        print_task(i)
        print()


#   TODO сделать так, чтобы бот определял по id пользователя и выводил только его заявки

def my_status_id(id):
    for i in range(1, line):
        #   hk = 0
        if sheet.cell(row=i, column=3).value == id:
            #   hk = sheet.cell(row = ???)
            #   print(sheet.cell(row=(hk), column=line).value, end='\t')
            print_task(i)


def find_empty_line():
    global k
    k = 1
    while sheet.cell(column=1, row=k).value is not None:
        k = k + 1
    return k


wb = load_workbook(filename='proverochka.xlsx')
sheet = wb.active
line = find_empty_line()
main()
