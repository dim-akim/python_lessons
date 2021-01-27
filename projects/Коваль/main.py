from openpyxl import load_workbook as lw
from datetime import datetime

lesson_rows = (2, 5, 7, 9, 11, 13, 15)
homework_rows = (4, 6, 8, 10, 12, 14, 16)

weekday_column = {
    0: 2,  # Понедельник
    1: 5,  # Вторник
    2: 8,  # Среда
    3: 11,  # Четверг
    4: 14  # Пятница
}

file = lw('расписание.xlsx')
page = file.active


def main():
    print('__________________')
    print('Начало работы')

    while True:
        answer = listen()
        if answer == 'exit':
            break
        elif answer == 'Расписание':
            get_schedule()
        elif answer == 'ДЗ':
            get_homework()
        elif answer == 'Создать ДЗ':
            create_homework()
        else:
            print('Ошибка ввода')


def listen():
    return input('Введите команду. Варианты:\nРасписание\nДЗ\nСоздать ДЗ\n')


def get_schedule():
    """
    Функция Расписание.
    Варианты вывода: на сегодня, на завтра, на неделю
    :return:
    """

    today_weekday = datetime.weekday(datetime.today())
    today_weekday = weekday_column[today_weekday]

    answer = input('__________________\nВыберите один из вариантов:\nНа сегодня\nНа завтра\nНа неделю\n')
    print()
    if answer == 'На сегодня':
        return get_schedule_today(today_weekday)
    elif answer == 'На завтра':
        return get_schedule_tomorrow(today_weekday)
    elif answer == 'На неделю':
        return get_schedule_week()
    # elif answer == 'Назад':
    else:
        print('Ошибка ввода')


def get_schedule_today(weekday_column):
    """
    Выводит из файла расписание.xlsx расписание на текущий день
    :return:
    """
    print(page.cell(row=1, column=weekday_column).value)
    print('__________________')
    for row in lesson_rows:
        time = page.cell(row=row, column=1).value
        print(time, '\t', page.cell(row=row, column=weekday_column).value, end='\t\t')
        for i in range(1, 3):
            if page.cell(row=row, column=weekday_column + i).value is not None:
                print(page.cell(row=row, column=weekday_column + i).value, end='\t\t')
        print()  # Переход на новую строку.
    print('__________________')


def get_schedule_tomorrow(weekday_colummn):
    """выводит расписание на след день
    имполтзуется та же функция, что и при выводе на сегодня
    """
    if weekday_colummn == 14:
        weekday_colummn = -1
        # вывод расписания на понедельник при обращении к боту в пятницу
    get_schedule_today(weekday_colummn + 3)


def get_schedule_week():
    print('__________________')
    for day in weekday_column:
        get_schedule_today(weekday_column[day])
    print('__________________')


def get_homework():
    today_weekday = datetime.weekday(datetime.today())
    today_weekday = weekday_column[today_weekday]

    answer = input('__________________\nВыберите один из вариантов:\nНа сегодня\nНа завтра\nНа неделю\n')
    print()
    if answer == 'На сегодня':
        return get_homework_today(today_weekday)
    elif answer == 'На завтра':
        return get_homework_tomorrow(today_weekday)
    elif answer == 'На неделю':
        return get_homework_week()
    # elif answer == 'Назад':
    else:
        print('Ошибка ввода')


def get_homework_today(weekday_column):
    print(page.cell(row=1, column=weekday_column).value)
    print('__________________')
    for row in homework_rows:
        print(page.cell(row=row-1, column=weekday_column).value, end='\t\t')
        print(page.cell(row=row, column=weekday_column).value, end='\t\t')
        for i in range(1, 3):
            if page.cell(row=row, column=weekday_column + i).value is not None:
                print(page.cell(row=row, column=weekday_column + i).value, end='\t\t')
        print()  # Переход на новую строку.
    print('__________________')


def get_homework_tomorrow(weekday_column):
    get_homework_today(weekday_column + 3)
    print('__________________')


def get_homework_week():
    print('__________________')
    for day in weekday_column:
        get_homework_today(weekday_column[day])
    print('__________________')


def find_subject(lesson):
    # определяет координаты урока введенного пользователем
    result = []
    for column in weekday_column.values():
        for row in lesson_rows:
            if page.cell(row=row, column=column).value == lesson:
                result.append((column, row))
    if len(result) > 1:
        result = choice_lesson(result)
    else:
        result = result[0]

    return result


def get_lesson_name():
    # Пользователь вводит название предмета
    lesson = input('Введите название предмета: \n')
    lesson = lesson.lower()
    return lesson


def create_homework():
    lesson = get_lesson_name()
    content = get_content()
    column, row = find_subject(lesson)
    page.cell(row=row+1, column=column).value = content
    file.save('расписание.xlsx')


def get_content():
    # Пользователь вводит само задание
    content = input('Введите задание: \n')
    return content


def choice_lesson(result):
    # Пользователь выбирает из найденых уроков к какому прицепить дз
    print('Выберите урок, к которому прикрепить ДЗ')
    lesson_number = 1
    for col, row in result:
        day = page.cell(row=1, column=col).value
        time = page.cell(row=row, column=1).value
        print(f'{lesson_number}: {day} в {time}')
        lesson_number += 1
    number = int(input('Введите ваш выбор: '))
    return result[number - 1]


def write_homework(days):
    print()


if __name__ == '__main__':
    main()
